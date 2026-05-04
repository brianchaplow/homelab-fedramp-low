"""Render an OSCAL POA&M JSON into a FedRAMP POA&M xlsx.

The FedRAMP Rev 5 POA&M template has six sheets; this renderer writes
active items into the ``Open POA&M Items`` sheet starting at
:data:`POAM_DATA_START_ROW` -- row 8 -- which is the first blank row
after the header (row 5), guidance (row 6), and example (row 7) rows.
The Closed POA&M Items sheet and the top-of-document CSP/CSO/Impact
metadata cells (rows 2-3) are not touched by Plan 2; Plan 3 can add
those.

Column mapping was discovered via ``openpyxl`` read of row 5 in the
live template on 2026-04-09. Row 5 is the header row, not row 1 --
the template uses rows 1-4 for branded document header cells.

Severity translation: the internal ``Finding`` schema uses
``Critical/High/Medium/Low/Info`` but the FedRAMP POA&M template's
"Original Risk Rating" validation only accepts
``Low/Moderate/High/Critical`` -- ``Medium`` maps to ``Moderate`` and
``Info`` maps to ``Low``. The poam-state string is flattened to
template-native columns: ``False Positive`` → column 21 = "Yes".
``Vendor Dependency`` is always written as ``No`` because the
homelab is not vendor-dependent; a future phase can make this
data-driven.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pipelines.common.logging import get_logger


logger = get_logger(__name__)


POAM_SHEET_NAME: str = "Open POA&M Items"

# poam-state values that should NOT appear on the "Open POA&M Items"
# sheet. Items in these states have been adjudicated via Deviation
# Request and remain in the OSCAL POA&M JSON for audit traceability,
# but they are excluded from the active count auditors see in the
# headline xlsx. The full list of states lives in the FedRAMP
# OSCAL Rev 5 enum: Open, In Remediation, Risk Accepted, False
# Positive, Closed, Operational Requirement.
EXCLUDED_FROM_OPEN_SHEET: frozenset[str] = frozenset({
    "Risk Accepted",
    "False Positive",
    "Closed",
    "Operational Requirement",
})

# Data rows begin at row 8 in the Rev 5 template:
#   row 1-4: document header / CSP metadata rows
#   row 5:   column headers (POAM ID, Controls, Weakness Name, ...)
#   row 6:   per-column guidance text
#   row 7:   per-column examples
#   row 8+:  first available data row
POAM_DATA_START_ROW: int = 8


# Verified against FedRAMP-POAM-Template-Rev5.xlsx "Open POA&M Items"
# sheet row 5 on 2026-04-09.
COLUMN_MAP: dict[str, int] = {
    "poam-id": 2,                    # POAM ID
    "controls": 3,                   # Controls
    "title": 4,                      # Weakness Name
    "description": 5,                # Weakness Description
    "weakness-detector-source": 6,   # Weakness Detector Source
    "weakness-source-id": 7,         # Weakness Source Identifier
    "asset-identifier": 8,           # Asset Identifier
    "point-of-contact": 9,           # Point of Contact
    "resources-required": 10,        # Resources Required
    "remediation-plan": 11,          # Remediation Plan
    "original-detection": 12,        # Original Detection Date
    "scheduled-completion": 13,      # Scheduled Completion Date
    "status-date": 14,               # Status Date
    "vendor-dependency": 15,         # Vendor Dependency
    "vendor-check-in": 16,           # Last Vendor Check-in Date
    "vendor-product": 17,            # Vendor Dependent Product Name
    "original-risk-rating": 18,      # Original Risk Rating
    "adjusted-risk-rating": 19,      # Adjusted Risk Rating
    "risk-adjustment": 20,           # Risk Adjustment
    "false-positive": 21,            # False Positive
    "operational-requirement": 22,   # Operational Requirement
    "deviation-rationale": 23,       # Deviation Rationale
    "supporting-documents": 24,      # Supporting Documents
    "comments": 25,                  # Comments
    "bod-tracking": 26,              # BOD 22-01 tracking
    "bod-due-date": 27,              # BOD 22-01 Due Date
    "cve": 28,                       # CVE
}


# Severity strings used in the internal Finding schema → strings the
# FedRAMP POA&M template's "Original Risk Rating" dropdown accepts.
_SEVERITY_TO_FEDRAMP: dict[str, str] = {
    "Critical": "Critical",
    "High": "High",
    "Medium": "Moderate",
    "Low": "Low",
    "Info": "Low",
}


def _props_to_dict(props: list[dict[str, Any]]) -> dict[str, str]:
    return {p["name"]: p.get("value", "") for p in props}


def _fedramp_severity(internal_severity: str) -> str:
    return _SEVERITY_TO_FEDRAMP.get(internal_severity, internal_severity)


def render_poam_from_oscal(
    oscal_poam_path: Path,
    template_path: Path,
    output_path: Path,
) -> Path:
    """Populate the FedRAMP POA&M template from an OSCAL POA&M JSON.

    Args:
        oscal_poam_path: path to an OSCAL POA&M JSON (output of
            :func:`pipelines.build.oscal_poam.build_poam_from_defectdojo`).
        template_path: path to ``FedRAMP-POAM-Template-Rev5.xlsx``.
        output_path: where to write the rendered workbook. Parent
            directories are created if missing.

    Returns:
        The ``output_path`` that was written.
    """
    poam_data = json.loads(oscal_poam_path.read_text(encoding="utf-8"))
    items = poam_data["plan-of-action-and-milestones"]["poam-items"]

    wb = load_workbook(template_path)
    if POAM_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"POA&M template missing expected sheet '{POAM_SHEET_NAME}' -- "
            f"found {wb.sheetnames}"
        )
    ws = wb[POAM_SHEET_NAME]

    row_index = 0
    excluded = 0
    for item in items:
        props = _props_to_dict(item.get("props", []))

        weakness_id = props.get("weakness-id", "")
        poam_state = props.get("poam-state", "Open")
        severity = props.get("severity", "Low")
        fedramp_severity = _fedramp_severity(severity)
        is_false_positive = poam_state == "False Positive"

        # Items adjudicated via DR (Risk Accepted / False Positive /
        # Closed / Operational Requirement) stay in the OSCAL JSON for
        # audit traceability but are excluded from the headline xlsx.
        if poam_state in EXCLUDED_FROM_OPEN_SHEET:
            excluded += 1
            continue

        row = POAM_DATA_START_ROW + row_index
        row_index += 1

        def cell(name: str, value: Any) -> None:
            if value not in (None, ""):
                ws.cell(row=row, column=COLUMN_MAP[name], value=value)

        # POAM ID: prefix the weakness id to disambiguate from other sources
        cell("poam-id", f"MSS-{weakness_id}" if weakness_id else f"MSS-{row_index}")
        cell("controls", props.get("related-controls", ""))
        cell("title", item.get("title", ""))
        cell("description", item.get("description", ""))
        cell("weakness-detector-source", "Wazuh")
        cell("weakness-source-id", props.get("cve") or weakness_id or "")
        cell("asset-identifier", props.get("affected-host", ""))
        cell("point-of-contact", "MSS Operator")
        cell("original-detection", props.get("discovered-date", ""))
        cell("scheduled-completion", props.get("scheduled-completion", ""))
        cell("vendor-dependency", "No")
        cell("original-risk-rating", fedramp_severity)
        cell("risk-adjustment", "No")
        cell("false-positive", "Yes" if is_false_positive else "No")
        cell("cve", props.get("cve", ""))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(
        "wrote POA&M xlsx with %d open items (%d excluded by DR) starting at row %d -> %s",
        row_index,
        excluded,
        POAM_DATA_START_ROW,
        output_path,
    )
    return output_path
