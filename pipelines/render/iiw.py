"""Render an OSCAL component-definition into a FedRAMP IIW xlsx.

The FedRAMP Rev 5 Integrated Inventory Workbook template has three
sheets: ``INSTRUCTIONS``, ``Inventory``, and ``Record of Changes``. This
renderer writes data into the ``Inventory`` sheet starting at
:data:`IIW_TEMPLATE_DATA_START_ROW` -- row 13 -- which is the first row
*after* the 10-row example/guidance block (rows 3 through 12) that the
template itself marks for deletion before submission. The headers in
row 2 stay intact. Column A is the instructional "delete before
submission" column, which is also left alone.

The column layout was discovered by inspecting the live template on
2026-04-09 (``openpyxl`` read of row 2 values). The mapping from OSCAL
prop name → column index is captured in :data:`COLUMN_MAP` below so
other tests and the CLI can reference the same source of truth.

Boolean props (``is-virtual``, ``is-public``, ``authenticated-scan``,
``in-latest-scan``) are coerced to FedRAMP's ``Yes`` / ``No`` text
convention. The OS Name and Version column is combined from the
``os-name`` and ``os-version`` props because the template uses one
column for both. DNS Name falls back to the component ``title``
(hostname) when there is no explicit ``dns-name`` prop, since "which
hostname responds" is the most common use of that column.

**Submission scrub:** the rendered xlsx still contains the template's
example rows (3-12) and the instructional column A. Before submitting
to FedRAMP, delete rows 3 through 12 and delete column A as the
template's row-1 guidance instructs. Plan 2 leaves this as a manual
step for auditability; a future phase can add an automated scrub.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pipelines.common.logging import get_logger


logger = get_logger(__name__)


# Sheet the data goes into (the workbook has three; the other two are
# instructions and the record-of-changes log).
IIW_SHEET_NAME: str = "Inventory"


# Row where component data begins. Row 1 is the "DELETE COL A AND ROWS
# 3-12 BEFORE SUBMISSION" instruction. Row 2 is the headers. Rows 3-12
# are the template's example/guidance rows. Row 13 is the first blank
# row available for real data.
IIW_TEMPLATE_DATA_START_ROW: int = 13


# OSCAL prop name → 1-indexed column in the Inventory sheet.
# Verified against FedRAMP-IIW-Template-Rev5.xlsx row-2 headers
# (Inventory sheet) on 2026-04-09.
COLUMN_MAP: dict[str, int] = {
    "asset-id": 2,           # UNIQUE ASSET IDENTIFIER
    "ipv4-address": 3,       # IPv4 or IPv6 Address
    "is-virtual": 4,         # Virtual
    "is-public": 5,          # Public
    "dns-name": 6,           # DNS Name or URL
    "netbios-name": 7,       # NetBIOS Name
    "mac-address": 8,        # MAC Address
    "authenticated-scan": 9, # Authenticated Scan
    "baseline-config": 10,   # Baseline Configuration Name
    "os-name": 11,           # OS Name and Version (combined with os-version)
    "location": 12,          # Location
    "asset-type": 13,        # Asset Type
    "hardware-model": 14,    # Hardware Make/Model
    "in-latest-scan": 15,    # In Latest Scan
    "software-vendor": 16,   # Software/Database Vendor
    "software-name": 17,     # Software/Database Name & Version
    "patch-level": 18,       # Patch Level
    "diagram-label": 19,     # Diagram Label
    "comments": 20,          # Comments
    "asset-tag": 21,         # Serial #/Asset Tag#
    # col 22 = VLAN/Network ID        (not tracked in InventoryComponent)
    # col 23 = System Admin/Owner     (not tracked)
    # col 24 = Application Admin/Owner (not tracked)
    "function": 25,          # Function
    "end-of-life": 26,       # End-of-Life
}


# Props whose "true"/"false" value should render as "Yes"/"No".
_BOOLEAN_PROPS: frozenset[str] = frozenset(
    {"is-virtual", "is-public", "authenticated-scan", "in-latest-scan"}
)


def _props_to_dict(props: list[dict[str, Any]]) -> dict[str, str]:
    """Flatten an OSCAL props list into a ``{name: value}`` dict."""
    return {p["name"]: p.get("value", "") for p in props}


def _bool_to_yes_no(value: str) -> str:
    if value == "true":
        return "Yes"
    if value == "false":
        return "No"
    return value


def _os_name_and_version(props: dict[str, str]) -> str:
    name = props.get("os-name", "").strip()
    version = props.get("os-version", "").strip()
    return f"{name} {version}".strip()


def render_iiw_from_oscal(
    component_def_path: Path,
    template_path: Path,
    output_path: Path,
) -> Path:
    """Populate the FedRAMP IIW template with component-definition data.

    Args:
        component_def_path: path to an OSCAL ``component-definition`` JSON
            file (the output of :mod:`pipelines.build.oscal_component`).
        template_path: path to ``FedRAMP-IIW-Template-Rev5.xlsx``.
        output_path: where to write the rendered workbook. Parent
            directories are created if missing.

    Returns:
        The ``output_path`` that was written.
    """
    cd_data = json.loads(component_def_path.read_text(encoding="utf-8"))
    components = cd_data["component-definition"]["components"]

    wb = load_workbook(template_path)
    if IIW_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"IIW template missing expected sheet '{IIW_SHEET_NAME}' -- found {wb.sheetnames}"
        )
    ws = wb[IIW_SHEET_NAME]

    for index, component in enumerate(components):
        row = IIW_TEMPLATE_DATA_START_ROW + index
        props = _props_to_dict(component.get("props", []))
        title = component.get("title", "")

        # Combined OS Name + Version cell
        os_combined = _os_name_and_version(props)

        # DNS Name falls back to hostname when missing.
        dns_name = props.get("dns-name") or title

        for prop_name, col_idx in COLUMN_MAP.items():
            if prop_name == "os-name":
                value: str = os_combined
            elif prop_name == "dns-name":
                value = dns_name
            else:
                raw = props.get(prop_name, "")
                value = _bool_to_yes_no(raw) if prop_name in _BOOLEAN_PROPS else raw

            if value:
                ws.cell(row=row, column=col_idx, value=value)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(
        "wrote IIW xlsx with %d component rows starting at row %d -> %s",
        len(components),
        IIW_TEMPLATE_DATA_START_ROW,
        output_path,
    )
    return output_path
