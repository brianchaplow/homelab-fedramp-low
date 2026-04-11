"""Tests for pipelines.render.poam -- OSCAL POA&M → FedRAMP POA&M xlsx."""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from pipelines.render.poam import (
    COLUMN_MAP,
    POAM_DATA_START_ROW,
    POAM_SHEET_NAME,
    render_poam_from_oscal,
)


TEMPLATE = Path("templates/FedRAMP-POAM-Template-Rev5.xlsx")


@pytest.fixture
def sample_poam(tmp_path: Path) -> Path:
    poam = {
        "plan-of-action-and-milestones": {
            "uuid": "11111111-1111-1111-1111-111111111111",
            "metadata": {
                "title": "Test POA&M",
                "last-modified": "2026-04-07T12:00:00+00:00",
                "version": "0.1.0",
                "oscal-version": "1.1.2",
            },
            "system-id": {
                "id": "test-system",
                "identifier-type": "https://ietf.org/rfc/rfc4122",
            },
            "poam-items": [
                {
                    "uuid": "33333333-3333-3333-3333-333333333333",
                    "title": "openssl heap overflow",
                    "description": "A real CVE",
                    "props": [
                        {"name": "weakness-id", "value": "1234"},
                        {"name": "severity", "value": "High"},
                        {"name": "cve", "value": "CVE-2024-1234"},
                        {"name": "discovered-date", "value": "2026-04-07"},
                        {"name": "scheduled-completion", "value": "2026-05-07"},
                        {"name": "poam-state", "value": "Open"},
                        {"name": "affected-host", "value": "brisket"},
                        {"name": "related-controls", "value": "RA-5, SI-2"},
                    ],
                },
                {
                    "uuid": "44444444-4444-4444-4444-444444444444",
                    "title": "nginx request smuggling",
                    "description": "Medium finding",
                    "props": [
                        {"name": "weakness-id", "value": "2000"},
                        {"name": "severity", "value": "Medium"},
                        {"name": "discovered-date", "value": "2026-04-08"},
                        {"name": "scheduled-completion", "value": "2026-07-07"},
                        {"name": "poam-state", "value": "Open"},
                        {"name": "affected-host", "value": "haccp"},
                        {"name": "related-controls", "value": "RA-5"},
                    ],
                },
                {
                    "uuid": "55555555-5555-5555-5555-555555555555",
                    "title": "false-positive finding",
                    "description": "FP case",
                    "props": [
                        {"name": "weakness-id", "value": "3000"},
                        {"name": "severity", "value": "Low"},
                        {"name": "discovered-date", "value": "2026-04-01"},
                        {"name": "scheduled-completion", "value": "2026-09-28"},
                        {"name": "poam-state", "value": "False Positive"},
                        {"name": "affected-host", "value": "brisket"},
                    ],
                },
            ],
        }
    }
    p = tmp_path / "poam.json"
    p.write_text(json.dumps(poam), encoding="utf-8")
    return p


@pytest.mark.skipif(not TEMPLATE.exists(), reason="POA&M template missing")
def test_render_poam_creates_xlsx(sample_poam: Path, tmp_path: Path) -> None:
    out = tmp_path / "poam.xlsx"
    path = render_poam_from_oscal(sample_poam, TEMPLATE, out)
    assert path == out
    assert out.exists()


@pytest.mark.skipif(not TEMPLATE.exists(), reason="POA&M template missing")
def test_render_poam_writes_to_open_items_sheet(
    sample_poam: Path, tmp_path: Path
) -> None:
    out = tmp_path / "poam.xlsx"
    render_poam_from_oscal(sample_poam, TEMPLATE, out)
    wb = load_workbook(out)
    assert POAM_SHEET_NAME in wb.sheetnames
    ws = wb[POAM_SHEET_NAME]
    # First finding lands at POAM_DATA_START_ROW
    row = POAM_DATA_START_ROW
    assert ws.cell(row=row, column=COLUMN_MAP["title"]).value == "openssl heap overflow"
    assert ws.cell(row=row, column=COLUMN_MAP["cve"]).value == "CVE-2024-1234"


@pytest.mark.skipif(not TEMPLATE.exists(), reason="POA&M template missing")
def test_render_poam_severity_medium_becomes_moderate(
    sample_poam: Path, tmp_path: Path
) -> None:
    """FedRAMP POA&M template uses 'Moderate' -- map Medium → Moderate."""
    out = tmp_path / "poam.xlsx"
    render_poam_from_oscal(sample_poam, TEMPLATE, out)
    ws = load_workbook(out)[POAM_SHEET_NAME]
    row2 = POAM_DATA_START_ROW + 1
    assert ws.cell(row=row2, column=COLUMN_MAP["original-risk-rating"]).value == "Moderate"


@pytest.mark.skipif(not TEMPLATE.exists(), reason="POA&M template missing")
def test_render_poam_false_positive_flag(sample_poam: Path, tmp_path: Path) -> None:
    """poam-state 'False Positive' sets the False Positive column to 'Yes'."""
    out = tmp_path / "poam.xlsx"
    render_poam_from_oscal(sample_poam, TEMPLATE, out)
    ws = load_workbook(out)[POAM_SHEET_NAME]
    row3 = POAM_DATA_START_ROW + 2  # false-positive finding
    assert ws.cell(row=row3, column=COLUMN_MAP["false-positive"]).value == "Yes"
    assert (
        ws.cell(row=POAM_DATA_START_ROW, column=COLUMN_MAP["false-positive"]).value
        == "No"
    )


@pytest.mark.skipif(not TEMPLATE.exists(), reason="POA&M template missing")
def test_render_poam_populates_controls_and_asset(
    sample_poam: Path, tmp_path: Path
) -> None:
    out = tmp_path / "poam.xlsx"
    render_poam_from_oscal(sample_poam, TEMPLATE, out)
    ws = load_workbook(out)[POAM_SHEET_NAME]
    row = POAM_DATA_START_ROW
    assert ws.cell(row=row, column=COLUMN_MAP["controls"]).value == "RA-5, SI-2"
    assert ws.cell(row=row, column=COLUMN_MAP["asset-identifier"]).value == "brisket"
    assert (
        ws.cell(row=row, column=COLUMN_MAP["weakness-detector-source"]).value
        == "Wazuh"
    )


@pytest.mark.skipif(not TEMPLATE.exists(), reason="POA&M template missing")
def test_render_poam_populates_dates(sample_poam: Path, tmp_path: Path) -> None:
    out = tmp_path / "poam.xlsx"
    render_poam_from_oscal(sample_poam, TEMPLATE, out)
    ws = load_workbook(out)[POAM_SHEET_NAME]
    row = POAM_DATA_START_ROW
    assert ws.cell(row=row, column=COLUMN_MAP["original-detection"]).value == "2026-04-07"
    assert (
        ws.cell(row=row, column=COLUMN_MAP["scheduled-completion"]).value == "2026-05-07"
    )


@pytest.mark.skipif(not TEMPLATE.exists(), reason="POA&M template missing")
def test_render_poam_preserves_template_headers(
    sample_poam: Path, tmp_path: Path
) -> None:
    out = tmp_path / "poam.xlsx"
    render_poam_from_oscal(sample_poam, TEMPLATE, out)
    ws = load_workbook(out)[POAM_SHEET_NAME]
    assert ws.cell(row=5, column=2).value == "POAM ID"
    assert ws.cell(row=5, column=4).value == "Weakness Name"
    assert ws.cell(row=5, column=28).value == "CVE"


def test_column_map_has_required_keys() -> None:
    required = {
        "poam-id",
        "controls",
        "title",
        "description",
        "weakness-detector-source",
        "asset-identifier",
        "original-detection",
        "scheduled-completion",
        "original-risk-rating",
        "false-positive",
        "cve",
    }
    assert required.issubset(COLUMN_MAP.keys())
