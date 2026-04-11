"""Tests for pipelines.render.iiw -- OSCAL component-definition → FedRAMP IIW xlsx."""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from pipelines.render.iiw import (
    IIW_SHEET_NAME,
    IIW_TEMPLATE_DATA_START_ROW,
    COLUMN_MAP,
    render_iiw_from_oscal,
)


TEMPLATE = Path("templates/FedRAMP-IIW-Template-Rev5.xlsx")


@pytest.fixture
def sample_component_def(tmp_path: Path) -> Path:
    cd = {
        "component-definition": {
            "uuid": "11111111-1111-1111-1111-111111111111",
            "metadata": {
                "title": "Test MSS",
                "last-modified": "2026-04-07T12:00:00+00:00",
                "version": "0.1.0",
                "oscal-version": "1.1.2",
            },
            "components": [
                {
                    "uuid": "33333333-3333-3333-3333-333333333333",
                    "type": "hardware",
                    "title": "brisket",
                    "description": "SIEM Manager",
                    "purpose": "SIEM Manager",
                    "props": [
                        {"name": "asset-id", "value": "MSS-HOST-BRISKET"},
                        {"name": "ipv4-address", "value": "10.10.20.30"},
                        {"name": "asset-type", "value": "Server"},
                        {"name": "is-virtual", "value": "false"},
                        {"name": "is-public", "value": "false"},
                        {"name": "authenticated-scan", "value": "true"},
                        {"name": "in-latest-scan", "value": "true"},
                        {"name": "os-name", "value": "Ubuntu"},
                        {"name": "os-version", "value": "24.04.4 LTS"},
                        {"name": "hardware-model", "value": "Lenovo ThinkStation"},
                        {"name": "patch-level", "value": "6.8.0-107-generic"},
                        {"name": "software-vendor", "value": "Ubuntu"},
                        {"name": "software-name", "value": "Ubuntu 24.04.4 LTS"},
                        {"name": "location", "value": "On-prem rack"},
                        {"name": "function", "value": "SIEM Manager"},
                        {"name": "diagram-label", "value": "S1"},
                        {"name": "asset-tag", "value": "HOMELAB-001"},
                        {"name": "end-of-life", "value": "2029-04"},
                        {"name": "comments", "value": "Core SIEM host"},
                        {"name": "boundary", "value": "in"},
                    ],
                },
                {
                    "uuid": "44444444-4444-4444-4444-444444444444",
                    "type": "network",
                    "title": "mokerlink",
                    "description": "Microsegmentation Switch",
                    "purpose": "Microsegmentation Switch",
                    "props": [
                        {"name": "asset-id", "value": "MSS-NON-MOKERLINK"},
                        {"name": "ipv4-address", "value": "10.10.10.2"},
                        {"name": "asset-type", "value": "Network"},
                        {"name": "is-virtual", "value": "false"},
                        {"name": "authenticated-scan", "value": "false"},
                        {"name": "hardware-model", "value": "MokerLink 10G08410GSM"},
                        {"name": "function", "value": "L2/L3 microsegmentation"},
                        {"name": "boundary", "value": "in"},
                    ],
                },
            ],
        }
    }
    p = tmp_path / "cd.json"
    p.write_text(json.dumps(cd), encoding="utf-8")
    return p


@pytest.mark.skipif(not TEMPLATE.exists(), reason="IIW template missing")
def test_render_iiw_creates_xlsx(sample_component_def: Path, tmp_path: Path) -> None:
    output = tmp_path / "iiw.xlsx"
    path = render_iiw_from_oscal(
        component_def_path=sample_component_def,
        template_path=TEMPLATE,
        output_path=output,
    )
    assert path == output
    assert output.exists()


@pytest.mark.skipif(not TEMPLATE.exists(), reason="IIW template missing")
def test_render_iiw_writes_to_inventory_sheet(
    sample_component_def: Path, tmp_path: Path
) -> None:
    output = tmp_path / "iiw.xlsx"
    render_iiw_from_oscal(sample_component_def, TEMPLATE, output)
    wb = load_workbook(output)
    assert IIW_SHEET_NAME in wb.sheetnames
    ws = wb[IIW_SHEET_NAME]
    # Data starts at IIW_TEMPLATE_DATA_START_ROW -- find brisket there
    row = IIW_TEMPLATE_DATA_START_ROW
    asset_id = ws.cell(row=row, column=COLUMN_MAP["asset-id"]).value
    assert asset_id == "MSS-HOST-BRISKET"


@pytest.mark.skipif(not TEMPLATE.exists(), reason="IIW template missing")
def test_render_iiw_populates_all_mapped_columns_for_brisket(
    sample_component_def: Path, tmp_path: Path
) -> None:
    output = tmp_path / "iiw.xlsx"
    render_iiw_from_oscal(sample_component_def, TEMPLATE, output)
    ws = load_workbook(output)[IIW_SHEET_NAME]
    row = IIW_TEMPLATE_DATA_START_ROW  # brisket is first component

    def cell(prop: str) -> object:
        return ws.cell(row=row, column=COLUMN_MAP[prop]).value

    assert cell("asset-id") == "MSS-HOST-BRISKET"
    assert cell("ipv4-address") == "10.10.20.30"
    # Boolean coercion to Yes/No
    assert cell("is-virtual") == "No"
    assert cell("is-public") == "No"
    assert cell("authenticated-scan") == "Yes"
    assert cell("in-latest-scan") == "Yes"
    # Combined OS name+version
    assert cell("os-name") == "Ubuntu 24.04.4 LTS"
    assert cell("asset-type") == "Server"
    assert cell("hardware-model") == "Lenovo ThinkStation"
    assert cell("patch-level") == "6.8.0-107-generic"
    assert cell("software-vendor") == "Ubuntu"
    assert cell("software-name") == "Ubuntu 24.04.4 LTS"
    assert cell("location") == "On-prem rack"
    assert cell("function") == "SIEM Manager"
    assert cell("diagram-label") == "S1"
    assert cell("asset-tag") == "HOMELAB-001"
    assert cell("end-of-life") == "2029-04"
    assert cell("comments") == "Core SIEM host"


@pytest.mark.skipif(not TEMPLATE.exists(), reason="IIW template missing")
def test_render_iiw_dns_name_falls_back_to_hostname(
    sample_component_def: Path, tmp_path: Path
) -> None:
    """Components without a dns-name prop get their title (hostname)
    written into the DNS Name column as a reasonable fallback."""
    output = tmp_path / "iiw.xlsx"
    render_iiw_from_oscal(sample_component_def, TEMPLATE, output)
    ws = load_workbook(output)[IIW_SHEET_NAME]
    dns_col = COLUMN_MAP["dns-name"]
    assert ws.cell(row=IIW_TEMPLATE_DATA_START_ROW, column=dns_col).value == "brisket"
    assert (
        ws.cell(row=IIW_TEMPLATE_DATA_START_ROW + 1, column=dns_col).value
        == "mokerlink"
    )


@pytest.mark.skipif(not TEMPLATE.exists(), reason="IIW template missing")
def test_render_iiw_second_row_is_mokerlink(
    sample_component_def: Path, tmp_path: Path
) -> None:
    output = tmp_path / "iiw.xlsx"
    render_iiw_from_oscal(sample_component_def, TEMPLATE, output)
    ws = load_workbook(output)[IIW_SHEET_NAME]
    row2 = IIW_TEMPLATE_DATA_START_ROW + 1
    assert ws.cell(row=row2, column=COLUMN_MAP["asset-id"]).value == "MSS-NON-MOKERLINK"
    # authenticated-scan=false → "No"
    assert ws.cell(row=row2, column=COLUMN_MAP["authenticated-scan"]).value == "No"
    assert ws.cell(row=row2, column=COLUMN_MAP["asset-type"]).value == "Network"


@pytest.mark.skipif(not TEMPLATE.exists(), reason="IIW template missing")
def test_render_iiw_preserves_template_headers(
    sample_component_def: Path, tmp_path: Path
) -> None:
    """The template's row 2 headers must survive the render intact."""
    output = tmp_path / "iiw.xlsx"
    render_iiw_from_oscal(sample_component_def, TEMPLATE, output)
    ws = load_workbook(output)[IIW_SHEET_NAME]
    assert ws.cell(row=2, column=2).value == "UNIQUE ASSET IDENTIFIER"
    assert ws.cell(row=2, column=3).value == "IPv4 or IPv6 Address"
    assert ws.cell(row=2, column=25).value == "Function"


def test_column_map_has_expected_headers() -> None:
    """Sanity-check: the column map keys are the props the renderer
    knows how to translate."""
    required = {
        "asset-id",
        "ipv4-address",
        "is-virtual",
        "is-public",
        "dns-name",
        "netbios-name",
        "mac-address",
        "authenticated-scan",
        "baseline-config",
        "os-name",
        "location",
        "asset-type",
        "hardware-model",
        "in-latest-scan",
        "software-vendor",
        "software-name",
        "patch-level",
        "diagram-label",
        "comments",
        "asset-tag",
        "function",
        "end-of-life",
    }
    assert required.issubset(COLUMN_MAP.keys())
